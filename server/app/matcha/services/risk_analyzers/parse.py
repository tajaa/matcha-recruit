"""Ingestion + normalization — turn a raw upload (CSV / XLSX) or a document
extraction (10-K / P&L / loss run figures pulled by Gemini) into the single
``normalized`` model the analyzer packs consume.

Pure except for the optional openpyxl import (XLSX). CSV uses stdlib ``csv``.
Deterministic — a hallucinated figure can only enter through the document
EXTRACTION step (which the user confirms), never through this parser.
"""

from __future__ import annotations

import csv
import io

from .base import to_float
from .mapping import map_roles, infer_kind, guess_role

_MAX_ROWS = 100_000        # rows we will read
_STORED_POINTS = 5_000     # points per series we persist (downsample beyond this)
_MAX_COLS = 60
_NUMERIC_THRESHOLD = 0.6   # a column is numeric if ≥60% of non-blank cells parse


# --------------------------------------------------------------------------- #
# Raw rows → tabular parsed structure
# --------------------------------------------------------------------------- #

def _rows_from_csv(raw: bytes) -> list[list]:
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1", errors="replace")
    sample = text[:4096]
    delim = ","
    try:
        delim = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        pass
    rows: list[list] = []
    for i, row in enumerate(csv.reader(io.StringIO(text), delimiter=delim)):
        if i >= _MAX_ROWS:
            break
        rows.append(row)
    return rows


def _rows_from_xlsx(raw: bytes) -> list[list]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency present in prod
        raise RuntimeError("XLSX support requires openpyxl") from exc
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[list] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= _MAX_ROWS:
            break
        rows.append(["" if c is None else c for c in row])
    wb.close()
    return rows


def _is_numeric_cell(v) -> bool:
    return to_float(v) is not None


def _dedupe(labels: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for h in labels:
        base = str(h).strip() or "col"
        if base in seen:
            seen[base] += 1
            out.append(f"{base}_{seen[base]}")
        else:
            seen[base] = 0
            out.append(base)
    return out


def _looks_like_period_index(values: list) -> bool:
    """A first column is a period/index axis when it's mostly non-numeric
    labels (dates, 'Q1', fiscal-year strings) OR a strictly increasing run of
    integers (years / row sequence) — never a risk series."""
    filled = [c for c in values if str(c).strip()]
    if not filled:
        return False
    non_numeric = sum(1 for c in filled if not _is_numeric_cell(c))
    if non_numeric / len(filled) >= 0.5:
        return True
    nums_ = [to_float(c) for c in filled]
    if all(n is not None and float(n).is_integer() for n in nums_) and len(nums_) >= 3:
        if all(b > a for a, b in zip(nums_, nums_[1:])):  # strictly increasing ints
            return True
    return False


def _role_matches(names: list[str]) -> int:
    return sum(1 for n in names if guess_role(n))


def _build_series(names: list[str], columns_values: list[list]) -> dict[str, list]:
    """names[k] -> columns_values[k] (already a value list), keeping only the
    numeric ones, downsampled to the stored cap."""
    series: dict[str, list] = {}
    for name, raw_col in zip(names, columns_values):
        filled = [c for c in raw_col if str(c).strip() != ""]
        parsed = [to_float(c) for c in raw_col]
        parse_ok = sum(1 for p in parsed if p is not None)
        if filled and parse_ok / len(filled) >= _NUMERIC_THRESHOLD:
            col = parsed
            if len(col) > _STORED_POINTS:
                stride = len(col) // _STORED_POINTS + 1
                col = col[::stride]
            series[name] = col
    return series


def _tabular_from_rows(rows: list[list]) -> dict:
    warnings: list[str] = []
    rows = [r for r in rows if any(str(c).strip() for c in r)]
    if not rows:
        return {"columns": [], "series": {}, "periods": None,
                "row_count": 0, "column_count": 0, "truncated": False,
                "warnings": ["No data rows found."]}

    width = max(len(r) for r in rows)
    if width > _MAX_COLS:
        warnings.append(f"Only the first {_MAX_COLS} of {width} columns were read.")
        width = _MAX_COLS
    rows = [list(r[:width]) + [""] * (width - len(r[:width])) for r in rows]

    first = rows[0]
    has_header = any(str(c).strip() and not _is_numeric_cell(c) for c in first)
    if has_header:
        header = [str(c).strip() or f"col_{i + 1}" for i, c in enumerate(first)]
        body = rows[1:]
    else:
        header = [f"col_{i + 1}" for i in range(width)]
        body = rows
    if not body:
        return {"columns": [], "series": {}, "periods": None, "row_count": 0,
                "column_count": len(header), "truncated": False,
                "warnings": ["No data rows found."]}

    col0 = [r[0] for r in body]
    first_is_label = _looks_like_period_index(col0)

    # --- Orientation A: columns-as-series (time-series / wide tables) ---
    a_names = _dedupe(header[1:] if first_is_label else header)
    a_cols = [[r[j] for r in body] for j in (range(1, width) if first_is_label else range(width))]
    a_series = _build_series(a_names, a_cols)
    a_periods = [str(r[0]).strip() for r in body] if first_is_label else None
    a_score = _role_matches(list(a_series.keys()))

    # --- Orientation B: rows-as-series (line-items-in-rows P&L / loss run) ---
    # Only when the first column holds labels that map to roles (Revenue, COGS…).
    b_series: dict[str, list] = {}
    b_periods = None
    b_score = -1
    if first_is_label:
        row_labels = _dedupe([str(r[0]).strip() for r in body])
        row_role_hits = _role_matches(row_labels)
        if row_role_hits >= 2:
            b_periods = [str(h).strip() for h in header[1:]]
            b_cols = [[r[j] for j in range(1, width)] for r in body]
            b_series = _build_series(row_labels, b_cols)
            b_score = row_role_hits

    if b_score > a_score and b_series:
        series, periods = b_series, b_periods
        columns = [{"name": n, "numeric": True} for n in series]
        n_records = len(periods or [])
    else:
        series, periods = a_series, a_periods
        columns = [{"name": n, "numeric": n in series} for n in a_names]
        if first_is_label:
            columns = [{"name": header[0], "numeric": False, "role_hint": "period"}] + columns
        n_records = len(body)

    if not series:
        warnings.append("No numeric series detected — nothing to analyze.")

    return {
        "columns": columns,
        "series": series,
        "periods": periods,
        "row_count": n_records,
        "column_count": len(columns),
        "truncated": any(len(v) >= _STORED_POINTS for v in series.values()),
        "warnings": warnings,
    }


def parse_tabular(raw: bytes, source_kind: str) -> dict:
    """Parse CSV or XLSX bytes into the tabular ``parsed`` structure."""
    rows = _rows_from_xlsx(raw) if source_kind == "xlsx" else _rows_from_csv(raw)
    return _tabular_from_rows(rows)


# --------------------------------------------------------------------------- #
# Document extraction → parsed structure
# --------------------------------------------------------------------------- #

def parsed_from_extraction(extraction: dict) -> dict:
    """Convert a confirmed Gemini extraction (line-items × periods, with
    provenance) into the same tabular ``parsed`` shape. One line-item = one
    series; the values array aligns to ``periods``."""
    extraction = extraction or {}
    periods = [str(p) for p in (extraction.get("periods") or [])] or None
    series: dict[str, list] = {}
    columns: list[dict] = []
    provenance: dict[str, dict] = {}
    seen: dict[str, int] = {}
    for item in extraction.get("line_items") or []:
        if not isinstance(item, dict):
            continue
        label = str(item.get("label") or "").strip()
        if not label:
            continue
        if label in seen:
            seen[label] += 1
            label = f"{label} ({seen[label]})"
        else:
            seen[label] = 0
        vals = [to_float(v) for v in (item.get("values") or [])]
        if not any(v is not None for v in vals):
            continue
        series[label] = vals
        columns.append({"name": label, "numeric": True})
        provenance[label] = {"page": item.get("page"), "unit": item.get("unit")}
    return {
        "columns": columns,
        "series": series,
        "periods": periods,
        "row_count": len(periods or []),
        "column_count": len(series),
        "truncated": False,
        "warnings": list(extraction.get("notes") or []),
        "provenance": provenance,
    }


# --------------------------------------------------------------------------- #
# parsed → normalized (apply roles)
# --------------------------------------------------------------------------- #

def normalize(parsed: dict, *, source_kind: str, filename: str,
              roles_override: dict | None = None, kind_override: str | None = None) -> dict:
    """Build the unified ``normalized`` model, applying heuristic roles overlaid
    with any user override from the dataset's ``mapping`` JSONB."""
    series = parsed.get("series") or {}
    names = list(series.keys())
    roles = map_roles(names)
    if roles_override:
        for name, role in roles_override.items():
            if name in series:
                if role in (None, "", "none", "ignore"):
                    roles.pop(name, None)
                else:
                    roles[name] = role
    kind = kind_override or infer_kind(roles)
    meta = {
        "source_kind": source_kind,
        "filename": filename,
        "truncated": bool(parsed.get("truncated")),
        "warnings": list(parsed.get("warnings") or []),
    }
    if parsed.get("provenance"):
        meta["provenance"] = parsed["provenance"]
    return {
        "series": series,
        "periods": parsed.get("periods"),
        "roles": roles,
        "kind": kind,
        "meta": meta,
    }

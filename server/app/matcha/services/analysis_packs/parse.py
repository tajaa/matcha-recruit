"""Ingestion + normalization — turn a raw upload (CSV / XLSX) or a document
extraction (10-K / P&L / loss run figures pulled by Gemini) into the single
``normalized`` model the analyzer packs consume.

Pure except for the optional openpyxl import (XLSX). CSV uses stdlib ``csv``.
Deterministic — a hallucinated figure can only enter through the document
EXTRACTION step (which the user confirms), never through this parser.

Metrics are computed on the FULL parsed series; ``downsample_for_storage``
caps only what is persisted (and stamps ``meta.truncated`` + a warning at the
point the stride decision is made).
"""

from __future__ import annotations

import csv
import io
import re

from .base import to_float
from .mapping import map_roles, infer_kind, guess_role

_MAX_ROWS = 100_000        # rows we will read
_STORED_POINTS = 5_000     # points per series we PERSIST (metrics use the full series)
_MAX_COLS = 60
_NUMERIC_THRESHOLD = 0.6   # a column is numeric if ≥60% of non-blank cells parse

_PERIODISH_NAME = re.compile(r"\b(period|date|year|month|quarter|week|index|fy)\b", re.I)
_DATEISH = re.compile(r"^\d{4}[-/.]\d{1,2}([-/.]\d{1,2})?$|^\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}$")


# --------------------------------------------------------------------------- #
# Raw rows → tabular parsed structure
# --------------------------------------------------------------------------- #

def _rows_from_csv(raw: bytes) -> list[list]:
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1", errors="replace")
    sample = text[:4096]
    delim = ","
    try:
        delim = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        pass
    rows: list[list] = []
    for i, row in enumerate(csv.reader(io.StringIO(text), delimiter=delim)):
        if i >= _MAX_ROWS:
            break
        rows.append(row)
    return rows


def _rows_from_xlsx(raw: bytes) -> list[list]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency present in prod
        raise RuntimeError("XLSX support requires openpyxl") from exc
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[list] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= _MAX_ROWS:
            break
        rows.append(["" if c is None else c for c in row])
    wb.close()
    return rows


def _is_numeric_cell(v) -> bool:
    return to_float(v) is not None


def _dedupe(labels: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for h in labels:
        base = str(h).strip() or "col"
        if base in seen:
            seen[base] += 1
            out.append(f"{base}_{seen[base]}")
        else:
            seen[base] = 0
            out.append(base)
    return out


def _looks_like_period_index(values: list, name: str = "") -> bool:
    """A first column is a period/index axis when it's mostly non-numeric
    labels (dates, 'Q1', fiscal-year strings), OR an increasing integer run
    that also LOOKS like an axis — name-matched (period/date/year/index),
    year-ranged, or consecutive (+1 steps). A merely-increasing integer
    column (cumulative claims, running totals) stays a data series."""
    filled = [c for c in values if str(c).strip()]
    if not filled:
        return False
    non_numeric = sum(1 for c in filled if not _is_numeric_cell(c))
    if non_numeric / len(filled) >= 0.5:
        return True
    nums_ = [to_float(c) for c in filled]
    if not all(n is not None and float(n).is_integer() for n in nums_):
        return False
    if not all(b > a for a, b in zip(nums_, nums_[1:])):
        return False
    # An explicitly named axis (period/date/year/index) needs no minimum run —
    # the header says what it is. Unnamed integer runs need the ≥3 + shape
    # checks below so cumulative counts stay data series.
    if _PERIODISH_NAME.search(name or ""):
        return True
    if len(nums_) < 3:
        return False
    if all(1900 <= n <= 2100 for n in nums_):       # fiscal years
        return True
    if all(b - a == 1 for a, b in zip(nums_, nums_[1:])):  # row sequence
        return True
    return False


def _role_matches(names: list[str]) -> int:
    return sum(1 for n in names if guess_role(n))


def _build_series(names: list[str], columns_values: list[list]) -> dict[str, list]:
    """names[k] -> columns_values[k], keeping only the numeric columns.
    FULL resolution — persistence capping happens in downsample_for_storage."""
    series: dict[str, list] = {}
    for name, raw_col in zip(names, columns_values):
        filled = [c for c in raw_col if str(c).strip() != ""]
        parsed = [to_float(c) for c in raw_col]
        parse_ok = sum(1 for p in parsed if p is not None)
        if filled and parse_ok / len(filled) >= _NUMERIC_THRESHOLD:
            series[name] = parsed
    return series


def _tabular_from_rows(rows: list[list], force_orientation: str | None = None) -> dict:
    warnings: list[str] = []
    rows = [r for r in rows if any(str(c).strip() for c in r)]
    if not rows:
        return {"columns": [], "series": {}, "periods": None,
                "row_count": 0, "column_count": 0, "truncated": False,
                "warnings": ["No data rows found."]}

    width = max(len(r) for r in rows)
    if width > _MAX_COLS:
        warnings.append(f"Only the first {_MAX_COLS} of {width} columns were read.")
        width = _MAX_COLS
    rows = [list(r[:width]) + [""] * (width - len(r[:width])) for r in rows]

    # Header detection — two signals, both required to beat the failure modes:
    # (a) a non-numeric row-1 cell ATOP a numeric body column ("price" over
    #     numbers) — a headerless file whose first column is dates must NOT
    #     have its first data row eaten as a header;
    # (b) a label-column header ("metric,2021,2022,2023"): non-numeric,
    #     non-date-like first cell, numeric remaining header cells, atop a
    #     mostly non-numeric label column.
    first = rows[0]
    body_probe = rows[1:]
    has_header = False
    if body_probe:
        for j in range(width):
            cell = str(first[j]).strip()
            if not cell or _is_numeric_cell(cell):
                continue
            col_body = [r[j] for r in body_probe if str(r[j]).strip()]
            numeric_ok = sum(1 for c in col_body if _is_numeric_cell(c))
            if col_body and numeric_ok / len(col_body) >= _NUMERIC_THRESHOLD:
                has_header = True
                break
        if not has_header and width > 1:
            first_cell = str(first[0]).strip()
            others = [str(first[j]).strip() for j in range(1, width)]
            col0_body = [str(r[0]).strip() for r in body_probe if str(r[0]).strip()]
            col0_nonnum = sum(1 for c in col0_body if not _is_numeric_cell(c))
            if (first_cell and not _is_numeric_cell(first_cell)
                    and not _DATEISH.match(first_cell)
                    and any(others)
                    and all(_is_numeric_cell(o) for o in others if o)
                    and col0_body and col0_nonnum / len(col0_body) >= 0.5):
                has_header = True
    else:
        has_header = any(str(c).strip() and not _is_numeric_cell(c) for c in first)
    if has_header:
        header = [str(c).strip() or f"col_{i + 1}" for i, c in enumerate(first)]
        body = rows[1:]
    else:
        header = [f"col_{i + 1}" for i in range(width)]
        body = rows
    if not body:
        return {"columns": [], "series": {}, "periods": None, "row_count": 0,
                "column_count": len(header), "truncated": False,
                "warnings": ["No data rows found."]}

    col0 = [r[0] for r in body]
    first_is_label = _looks_like_period_index(col0, header[0])

    # --- Orientation A: columns-as-series (time-series / wide tables) ---
    a_names = _dedupe(header[1:] if first_is_label else header)
    a_cols = [[r[j] for r in body] for j in (range(1, width) if first_is_label else range(width))]
    a_series = _build_series(a_names, a_cols)
    a_periods = [str(r[0]).strip() for r in body] if first_is_label else None
    a_score = _role_matches(list(a_series.keys()))

    # --- Orientation B: rows-as-series (line-items-in-rows P&L / loss run) ---
    b_series: dict[str, list] = {}
    b_periods = None
    b_score = -1
    row_labels = _dedupe([str(r[0]).strip() or f"row_{i + 1}" for i, r in enumerate(body)])
    if first_is_label or force_orientation == "rows":
        row_role_hits = _role_matches(row_labels)
        if row_role_hits >= 2 or force_orientation == "rows":
            b_periods = [str(h).strip() for h in header[1:]]
            b_cols = [[r[j] for j in range(1, width)] for r in body]
            b_series = _build_series(row_labels, b_cols)
            b_score = row_role_hits

    if force_orientation == "rows":
        use_b = bool(b_series)
    elif force_orientation == "columns":
        use_b = False
    else:
        use_b = b_score > a_score and bool(b_series)

    if use_b:
        series, periods = b_series, b_periods
        columns = [{"name": n, "numeric": True} for n in series]
        n_records = len(periods or [])
        warnings.append("Rows were interpreted as series (line-items × period columns).")
    else:
        series, periods = a_series, a_periods
        columns = [{"name": n, "numeric": n in series} for n in a_names]
        if first_is_label:
            columns = [{"name": header[0], "numeric": False, "role_hint": "period"}] + columns
            warnings.append(f"Column '{header[0]}' was used as the period axis.")
        n_records = len(body)

    if not series:
        warnings.append("No numeric series detected — nothing to analyze.")

    return {
        "columns": columns,
        "series": series,
        "periods": periods,
        "row_count": n_records,
        "column_count": len(columns),
        "truncated": False,  # full resolution here; downsample_for_storage stamps truncation
        "warnings": warnings,
    }


def parse_tabular(raw: bytes, source_kind: str, force_orientation: str | None = None) -> dict:
    """Parse CSV or XLSX bytes into the tabular ``parsed`` structure at full
    resolution. ``force_orientation`` ('columns'|'rows') overrides the
    orientation heuristic — the user-facing escape hatch for misdetected
    layouts."""
    rows = _rows_from_xlsx(raw) if source_kind == "xlsx" else _rows_from_csv(raw)
    return _tabular_from_rows(rows, force_orientation)


def downsample_for_storage(normalized: dict) -> dict:
    """Cap each series to ``_STORED_POINTS`` for persistence (ceil-division
    stride) AFTER metrics have been computed on the full series. ``periods``
    are strided identically so labels stay aligned with the stored points.
    Stamps ``meta.truncated`` and a disclosure warning exactly when data was
    dropped."""
    series = normalized.get("series") or {}
    capped: dict[str, list] = {}
    dropped_any = False
    for name, values in series.items():
        n = len(values or [])
        if n > _STORED_POINTS:
            stride = -(-n // _STORED_POINTS)  # ceil division
            capped[name] = values[::stride]
            dropped_any = True
        else:
            capped[name] = values
    if not dropped_any:
        return normalized
    periods = normalized.get("periods")
    if periods and len(periods) > _STORED_POINTS:
        stride = -(-len(periods) // _STORED_POINTS)
        periods = periods[::stride]
    meta = dict(normalized.get("meta") or {})
    meta["truncated"] = True
    meta["warnings"] = list(meta.get("warnings") or []) + [
        "Series were downsampled for storage; metrics were computed on the full data."
    ]
    return {**normalized, "series": capped, "periods": periods, "meta": meta}


# --------------------------------------------------------------------------- #
# Document extraction → parsed structure
# --------------------------------------------------------------------------- #

def parsed_from_extraction(extraction: dict) -> dict:
    """Convert a confirmed Gemini extraction (line-items × periods, with
    provenance) into the same tabular ``parsed`` shape. One line-item = one
    series; the values array aligns to ``periods``."""
    extraction = extraction or {}
    periods = [str(p) for p in (extraction.get("periods") or [])] or None
    series: dict[str, list] = {}
    columns: list[dict] = []
    provenance: dict[str, dict] = {}
    seen: dict[str, int] = {}
    for item in extraction.get("line_items") or []:
        if not isinstance(item, dict):
            continue
        label = str(item.get("label") or "").strip()
        if not label:
            continue
        if label in seen:
            seen[label] += 1
            label = f"{label} ({seen[label]})"
        else:
            seen[label] = 0
        vals = [to_float(v) for v in (item.get("values") or [])]
        if not any(v is not None for v in vals):
            continue
        series[label] = vals
        columns.append({"name": label, "numeric": True})
        provenance[label] = {"page": item.get("page"), "unit": item.get("unit")}
    return {
        "columns": columns,
        "series": series,
        "periods": periods,
        "row_count": len(periods or []),
        "column_count": len(series),
        "truncated": False,
        "warnings": list(extraction.get("notes") or []),
        "provenance": provenance,
    }


# --------------------------------------------------------------------------- #
# parsed → normalized (apply roles)
# --------------------------------------------------------------------------- #

def normalize(parsed: dict, *, source_kind: str, filename: str,
              roles_override: dict | None = None, kind_override: str | None = None) -> dict:
    """Build the unified ``normalized`` model, applying heuristic roles overlaid
    with any user override from the dataset's ``mapping`` JSONB."""
    series = parsed.get("series") or {}
    names = list(series.keys())
    roles = map_roles(names)
    if roles_override:
        for name, role in roles_override.items():
            if name in series:
                if role in (None, "", "none", "ignore"):
                    roles.pop(name, None)
                else:
                    roles[name] = role
    kind = kind_override or infer_kind(roles)
    meta = {
        "source_kind": source_kind,
        "filename": filename,
        "truncated": bool(parsed.get("truncated")),
        "warnings": list(parsed.get("warnings") or []),
    }
    if parsed.get("provenance"):
        meta["provenance"] = parsed["provenance"]
    return {
        "series": series,
        # Explicit column list so slim loads (which drop the heavy `series`
        # values) still know the series names for the mapping UI.
        "columns": names,
        "periods": parsed.get("periods"),
        "roles": roles,
        "kind": kind,
        "meta": meta,
    }
